from datetime import timedelta
from django.contrib import admin
from django.db.models import JSONField
from django.shortcuts import redirect
from django.utils import timezone
from django_json_widget.widgets import JSONEditorWidget
from django_object_actions import DjangoObjectActions, action

from apps.vendor.tasks import contact_report, get_latest_sdn_date, vendor_report

from .models import (
    # TestSDNServices,
    BusinessUnit,
    Category,
    Department,
    OFACSDNResult,
    OFACSpeciallyDesignatedNational,
    OFACSpeciallyDesignatedNationalAddress,
    OFACSpeciallyDesignatedNationalAlternate,
    OrganizationSetting,
    Process,
    ScheduledJobLogger,
    BulkImport,
    TestEmailService,
    SiteFeatures,
    BusinessProcess,
    Location,
)
from django.shortcuts import redirect
from django.utils import timezone
from datetime import timedelta
from django.urls import re_path
from organizations.models import Organization
from apps.vendor.models import Contact, ContactRole, Phone, Vendor, PhoneType, get_risk, get_status
from openpyxl import load_workbook
from apps.authentication.models import User as AuthUser

from .tasks import fetch_weekly_ofac_list

# Register your models here.

admin.site.register(Location)
admin.site.register(Category)
admin.site.register(OrganizationSetting)
admin.site.register(BusinessUnit)
admin.site.register(Department)
admin.site.register(Process)
admin.site.register(BusinessProcess)

admin.site.register(SiteFeatures)


class OFACAddressInline(admin.TabularInline):
    model = OFACSpeciallyDesignatedNationalAddress

    def has_add_permission(self, *args) -> bool:
        return False

    def has_change_permission(self, *args) -> bool:
        return False

    def has_delete_permission(self, *args) -> bool:
        return False


class OFACAlternateInline(admin.TabularInline):
    model = OFACSpeciallyDesignatedNationalAlternate

    def has_add_permission(self, *args) -> bool:
        return False

    def has_change_permission(self, *args) -> bool:
        return False

    def has_delete_permission(self, *args) -> bool:
        return False


class OFACSDNAdmin(DjangoObjectActions, admin.ModelAdmin):
    list_filter = ("created_at", "sdn_type")
    list_display = ("id", "name", "sdn_type", "program")
    search_fields = ("name",)
    inlines = (OFACAddressInline, OFACAlternateInline)
    changelist_actions = ("import_ofac_data", "run_contact_search", "run_vendor_search", "delete_latest_ofac_data")

    def has_add_permission(self, *args) -> bool:
        return False

    def has_change_permission(self, *args) -> bool:
        return False

    def has_delete_permission(self, *args) -> bool:
        return False

    @action(label="Import", description="Import latest OFAC data.")
    def import_ofac_data(self, request, queryset):
        fetch_weekly_ofac_list()

    @action(label="Delete Latest", description="Delete latest OFAC data.")
    def delete_latest_ofac_data(self, request, queryset):
        latest_date = get_latest_sdn_date()
        OFACSpeciallyDesignatedNational.objects.filter(created_at=latest_date).delete()

    @action(label="Scan Contacts", description="Run all contacts against the latest OFAC data.")
    def run_contact_search(self, request, queryset):
        contact_report()

    @action(label="Scan 3rd Parties", description="Run all vendors against the latest OFAC data.")
    def run_vendor_search(self, request, queryset):
        vendor_report()


class OFACSDNAddressAdmin(admin.ModelAdmin):
    list_filter = ("created_at",)
    list_display = ("id", "sdn", "address", "country")
    search_fields = ("address",)

    def has_add_permission(self, *args) -> bool:
        return False

    def has_change_permission(self, *args) -> bool:
        return False

    def has_delete_permission(self, *args) -> bool:
        return False


class OFACSDNAlternateAdmin(admin.ModelAdmin):
    list_filter = ("created_at",)
    list_display = ("id", "sdn", "alt_name", "alt_type")
    search_fields = ("alt_name",)

    def has_add_permission(self, *args) -> bool:
        return False

    def has_change_permission(self, *args) -> bool:
        return False

    def has_delete_permission(self, *args) -> bool:
        return False


class OFACSDNResultAdmin(admin.ModelAdmin):
    list_display = ("id", "vendor", "contact", "total_address", "total_alias")
    search_fields = ("vendor__name", "vendor__legal_name", "contact__first_name", "contact__last_name")

    formfield_overrides = {JSONField: {"widget": JSONEditorWidget}}


admin.site.register(OFACSpeciallyDesignatedNational, OFACSDNAdmin)
admin.site.register(OFACSpeciallyDesignatedNationalAddress, OFACSDNAddressAdmin)
admin.site.register(OFACSpeciallyDesignatedNationalAlternate, OFACSDNAlternateAdmin)
admin.site.register(OFACSDNResult, OFACSDNResultAdmin)


class TestEmailServiceAdmin(admin.ModelAdmin):
    def render_change_form(self, request, context, add=False, change=False, form_url="", obj=None):
        context.update(
            {
                "show_save": True,
                "show_save_and_continue": False,
                "show_save_and_add_another": False,
                "show_delete": True,
            }
        )
        return super().render_change_form(request, context, add, change, form_url, obj)


admin.site.register(TestEmailService, TestEmailServiceAdmin)


class LoggerAdmin(admin.ModelAdmin):
    change_list_template = "admin_templates/change_list.html"
    list_display = (
        "get_created_at",
        "api_url",
        "email",
    )
    list_filter = ("email", "created_at")
    search_fields = [
        "user__email",
    ]

    @admin.display(ordering="created_at", description="Created Date/Time")
    def get_created_at(self, obj):
        return obj.created_at.strftime("%m/%d/%Y %H:%M")

    def get_urls(self):
        urls = super(LoggerAdmin, self).get_urls()
        new_urls = [
            re_path(r"^delete-logger", self.admin_site.admin_view(self.delete_logger)),
        ]
        return new_urls + urls

    def delete_logger(self, request):
        today = timezone.now().date()
        today = today - timedelta(days=int(request.GET.get("days")))
        logger_object = ScheduledJobLogger.objects.filter(created_at__date__lte=today).delete()
        return redirect(".")


admin.site.register(ScheduledJobLogger, LoggerAdmin)


class BulkImportAdmin(admin.ModelAdmin):
    change_list_template = "admin_bulk_import/change_form_template.html"

    def changelist_view(self, request, extra_context=None):
        orgs = Organization.objects.all()
        extra_context = {"title": "Upload file to import vendor list", "orgs": orgs}
        return super(BulkImportAdmin, self).changelist_view(request, extra_context=extra_context)

    def get_urls(self):
        urls = super(BulkImportAdmin, self).get_urls()
        new_urls = [
            re_path(r"^bulk-import", self.admin_site.admin_view(self.bulk_import)),
        ]
        return new_urls + urls

    def get_business_process(self, unit, department, process):
        if unit:
            return f"{unit}-{department}-{process}"
        elif department:
            return f"{department}-{process}"
        else:
            return process

    def bulk_import(self, request):
        if request.method == "POST":
            file = request.FILES["file"]
            wb = load_workbook(filename=file.file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if not row[0].value:
                    break
                owner = None
                if row[2].value:
                    owner_object = AuthUser.objects.filter(email=row[2].value)
                    if owner_object:
                        owner = owner_object.first()
                vendor = Vendor.objects.create(
                    org_id=request.POST["organization"],
                    name=row[0].value,
                    legal_name=row[1].value,
                    owner=owner,
                    category=row[3].value,
                    org_business_process=self.get_business_process(row[4].value, row[5].value, row[6].value),
                    is_offshore=True if row[7].value == "Y" else False,
                    inherent_risk=get_risk(row[9].value),
                    residual_risk=get_risk(row[11].value),
                    critical=True if row[12].value == "Y" else False,
                    status=get_status(row[13].value),
                )
                if row[14].value or row[15].value or row[16].value or row[17].value or row[19].value:
                    contact = Contact.objects.create(
                        vendor=vendor,
                        first_name="Primary",
                        last_name="Location",
                        email=row[21].value,
                        role=ContactRole.PRIMARY_BUSINESS_ADDRESS,
                        line_1=row[14].value,
                        line_2=row[15].value,
                        city=row[16].value,
                        state=row[17].value,
                        zip_code=row[19].value,
                    )
                if row[20].value:
                    Phone.objects.create(
                        contact=contact, number=row[20].value, is_preferred=False, type=PhoneType.MAIN
                    )
            return redirect(".")


admin.site.register(BulkImport, BulkImportAdmin)


# class TestSDNServiceAdmin(admin.ModelAdmin):
#     def render_change_form(self, request, context, add=False, change=False, form_url="", obj=None):
#         context.update(
#             {
#                 "show_save": True,
#                 "show_save_and_continue": False,
#                 "show_save_and_add_another": False,
#                 "show_delete": True,
#             }
#         )
#         return super().render_change_form(request, context, add, change, form_url, obj)


# admin.site.register(TestSDNServices, TestSDNServiceAdmin)
